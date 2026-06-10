from typing import Any, cast

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import generics, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response

from ..constants import Errors
from ..models import OlympiadParticipation
from ..queries.reports import (
    report_education,
    report_municipality,
    report_rating,
    report_stage_year,
    report_subject,
    report_winners,
)
from ..serializers.reports import ReportFiltersSerializer

# =================================================
# Конфигурация отчётов
# =================================================

REPORTS: dict[str, dict[str, Any]] = {
    "stage_year": {
        "title": "Сводка по этапам и годам",
        "description": "Количество участий, участников, победителей и призёров в разрезе этапов и годов.",
        "filename": "отчёт_этапы_годы.xlsx",
        "query_fn": report_stage_year,
        "columns": [
            ("year", "Год", 12),
            ("stage", "Этап", 18),
            ("participations", "Участий", 15),
            ("participants", "Участников", 15),
            ("winners", "Победителей", 15),
            ("prize_winners", "Призёров", 15),
        ],
    },
    "municipality": {
        "title": "Отчёт по муниципалитетам",
        "description": "Статистика участия по муниципалитетам: участники, победители, призёры, количество учреждений.",
        "filename": "отчёт_муниципалитеты.xlsx",
        "query_fn": report_municipality,
        "columns": [
            ("municipality_name", "Муниципалитет", 30),
            ("participations", "Участий", 15),
            ("participants", "Участников", 15),
            ("winners", "Победителей", 15),
            ("prize_winners", "Призёров", 15),
            ("educations", "Учреждений", 15),
        ],
    },
    "subject": {
        "title": "Отчёт по предметам",
        "description": "Сводные показатели участия по каждому предмету олимпиады.",
        "filename": "отчёт_предметы.xlsx",
        "query_fn": report_subject,
        "columns": [
            ("subject_name", "Предмет", 35),
            ("participations", "Участий", 15),
            ("participants", "Участников", 15),
            ("winners", "Победителей", 15),
            ("prize_winners", "Призёров", 15),
        ],
    },
    "education": {
        "title": "Отчёт по учреждениям",
        "description": "Детализация по образовательным учреждениям: участники, победители, призёры.",
        "filename": "отчёт_учреждения.xlsx",
        "query_fn": report_education,
        "columns": [
            ("municipality_name", "Муниципалитет", 30),
            ("education_name", "Учреждение", 50),
            ("participations", "Участий", 15),
            ("participants", "Участников", 15),
            ("winners", "Победителей", 15),
            ("prize_winners", "Призёров", 15),
        ],
    },
    "winners": {
        "title": "Победители и призёры",
        "description": "Список обучающихся со статусами «победитель» и «призёр» с полными данными.",
        "filename": "отчёт_победители_призёры.xlsx",
        "query_fn": report_winners,
        "columns": [
            ("full_name", "ФИО", 35),
            ("participant__birth_date", "Дата рождения", 16),
            ("participant__gender", "Пол", 8),
            ("subject__full_name", "Предмет", 30),
            ("class_field", "Класс", 10),
            ("stage", "Этап", 14),
            ("status", "Статус", 16),
            ("year", "Год", 10),
            ("education__full_name", "Учреждение", 50),
            ("education__municipality__name", "Муниципалитет", 30),
        ],
    },
    "rating": {
        "title": "Рейтинг обучающихся",
        "description": "Рейтинг по суммарным баллам: победитель = 5, призёр = 3, участник = 1.",
        "filename": "отчёт_рейтинг.xlsx",
        "query_fn": report_rating,
        "columns": [
            ("full_name", "ФИО", 35),
            ("participant_birth_date", "Дата рождения", 16),
            ("participant_gender", "Пол", 8),
            ("municipality_name", "Муниципалитет", 30),
            ("education_name", "Учреждение", 50),
            ("total_participations", "Участий", 12),
            ("total_winners", "Побед", 10),
            ("total_prize_winners", "Призовых", 10),
            ("total_participants", "Участий (статус)", 15),
            ("rating", "Рейтинг", 12),
        ],
    },
}


# =================================================
# Excel-стили
# =================================================

HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="435AAA", end_color="435AAA", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
TITLE_FONT = Font(name="Arial", bold=True, size=14)
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="666666")
TOTAL_FONT = Font(name="Arial", bold=True, size=10)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# =================================================
# Фильтрация queryset
# =================================================

FILTER_MAP = {
    "year": "year__in",
    "stage": "stage__in",
    "status": "status__in",
    "subject_id": "subject_id__in",
    "municipality_id": "education__municipality_id__in",
    "education_id": "education_id__in",
}


def _apply_filters(filters: dict[str, Any]):
    """Применяет фильтры к базовому queryset участий."""
    qs = OlympiadParticipation.objects.select_related(
        "participant",
        "education",
        "education__municipality",
        "subject",
    )

    for param, lookup in FILTER_MAP.items():
        values = filters.get(param)
        if values:
            qs = qs.filter(**{lookup: values})

    return qs


# =================================================
# Описание фильтров для Excel
# =================================================

FILTER_LABELS = {
    "year": ("Год", lambda v: ", ".join(str(x) for x in v)),
    "stage": ("Этап", lambda v: ", ".join(v)),
    "status": ("Статус", lambda v: ", ".join(v)),
    "subject_id": ("Предмет", lambda v: f"{len(v)} шт."),
    "municipality_id": ("Муниципалитет", lambda v: f"{len(v)} шт."),
    "education_id": ("Учреждение", lambda v: f"{len(v)} шт."),
}


def _describe_filters(filters: dict[str, Any]) -> str:
    """Формирует описание активных фильтров для вставки в Excel."""
    parts = []

    for key, (label, formatter) in FILTER_LABELS.items():
        values = filters.get(key)
        if values:
            parts.append(f"{label}: {formatter(values)}")

    return " | ".join(parts) if parts else ""


# =================================================
# Генерация Excel
# =================================================


def _generate_excel(report_key: str, filters: dict[str, Any]) -> HttpResponse:
    """Формирует Excel-файл по ключу отчёта и фильтрам."""
    config = REPORTS[report_key]
    columns = config["columns"]

    queryset = _apply_filters(filters)
    rows = config["query_fn"](queryset)

    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = config["title"][:31]

    col_count = len(columns)

    # Строка 1: заголовок отчёта
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=config["title"])
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Строка 2: описание фильтров
    filter_text = _describe_filters(filters)
    if filter_text:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        filter_cell = ws.cell(row=2, column=1, value=f"Фильтры: {filter_text}")
        filter_cell.font = SUBTITLE_FONT
        filter_cell.alignment = Alignment(horizontal="left")

    # Строка 3: пояснение для рейтинга
    if report_key == "rating":
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)
        rating_cell = ws.cell(
            row=3,
            column=1,
            value="Баллы: победитель = 5, призёр = 3, участник = 1",
        )
        rating_cell.font = SUBTITLE_FONT
        rating_cell.alignment = Alignment(horizontal="left")

    # Строка 4: шапка таблицы
    header_row = 4
    for col_idx, (_, label, width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Строки 5+: данные
    for row_idx, row_data in enumerate(rows, header_row + 1):
        for col_idx, (field, _, _) in enumerate(columns, 1):
            value = row_data.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGN
            cell.border = BORDER

    # Итого
    total_row = header_row + len(rows) + 1
    ws.cell(row=total_row, column=1, value=f"Итого записей: {len(rows)}")
    ws.cell(row=total_row, column=1).font = TOTAL_FONT

    ws.freeze_panes = f"A{header_row + 1}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{config["filename"]}"'
    response["Access-Control-Expose-Headers"] = "Content-Disposition"
    wb.save(response)

    return response


# =================================================
# Views
# =================================================


class ReportsListView(generics.GenericAPIView):
    """
    Список доступных отчётов.
    GET /vsosh/reports/
    """

    pagination_class = None
    permission_classes = [IsAuthenticated]

    def get(self, request: Request) -> Response:
        """Получить весь список"""
        reports = [
            {
                "key": key,
                "title": config["title"],
                "description": config["description"],
                "filename": config["filename"],
            }
            for key, config in REPORTS.items()
        ]

        return Response(reports)


class ReportDownloadView(generics.GenericAPIView):
    """
    Генерация и скачивание Excel-отчёта с фильтрами.
    GET /vsosh/reports/<report_key>/?year=2024&stage=РЭ
    """

    pagination_class = None
    permission_classes = [IsAuthenticated]
    serializer_class = ReportFiltersSerializer

    def get(self, request: Request, report_key: str):
        """Получить файл Excel-отчёта"""
        if report_key not in REPORTS:
            return Response(
                {"detail": Errors.REPORT_NOT_FOUND.format(report_key=report_key)},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = self.get_serializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)

        return _generate_excel(report_key, serializer.validated_data)
