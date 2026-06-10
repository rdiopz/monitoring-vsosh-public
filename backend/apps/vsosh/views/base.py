import logging
from datetime import date, datetime, timedelta
from typing import Any

import pandas as pd
from django.db import models, transaction
from django.db.models.functions import Lower
from django.forms.models import model_to_dict
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import MultiPartParser
from rest_framework.request import Request
from rest_framework.response import Response

from apps.common.utils import audit_log

from ..cache_service import VSOSHCacheService
from ..constants import Errors, Limits, Messages
from ..models import (
    EducationInstitution,
    Municipality,
    Participant,
    Subject,
)
from ..serializers.common import ExportRequestSerializer

logger = logging.getLogger(__name__)


# =============================================================================
# Вспомогательные функции
# =============================================================================


def _model_to_dict_with_fk_ids(
    instance: models.Model,
    exclude: set[str] | None = None,
) -> dict[str, Any]:
    """Конвертирует модель в dict, заменяя ForeignKey на их ID"""
    exclude = exclude or set()
    data = model_to_dict(instance, exclude=exclude)

    for field in instance._meta.fields:
        if field.is_relation and field.many_to_one:
            field_name = field.name
            if field_name in data:
                field_id = f"{field_name}_id"
                if hasattr(instance, field_id):
                    data[field_id] = getattr(instance, field_id)
                    del data[field_name]

    return data


def _get_old_values(request: Request) -> dict | None:
    """Получить сохранённые старые данные из request для аудита"""
    return getattr(request, "_audit_old_data", None)


def _get_export_new_values(request: Request, response_data: dict) -> dict | None:
    """Получить параметры экспорта для аудита"""
    return {
        "columns": request.query_params.getlist("columns"),
        "rows_count": getattr(request, "_export_rows_count", 0),
        "format": "xlsx",
    }


def _get_import_new_values(request: Request, response_data: dict) -> dict | None:
    """Получить параметры импорта для аудита"""
    return {
        "filename": request.FILES.get("file").name if request.FILES.get("file") else None,
        "created": response_data.get("created", 0),
        "updated": response_data.get("updated", 0),
        "errors": response_data.get("errors", 0),
    }


def parse_russian_date(date_value: Any) -> date:
    """Универсальный парсер дат с поддержкой русских форматов"""
    if date_value is None:
        raise ValueError(Errors.DATE_EMPTY)

    if isinstance(date_value, date):
        return date_value

    # Excel serial date (число дней с 30.12.1899)
    if isinstance(date_value, (int, float)):
        epoch = datetime(1899, 12, 30)
        return (epoch + timedelta(days=date_value)).date()

    if isinstance(date_value, str):
        date_str = date_value.strip()

        formats = [
            "%d.%m.%Y",  # 01.04.2026
            "%Y-%m-%d",  # 2026-04-01
            "%Y-%m-%d %H:%M:%S",  # 2026-04-01 00:00:00
            "%d.%m.%Y %H:%M:%S",  # 01.04.2026 00:00:00
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        raise ValueError(Errors.DATE_INVALID_FORMAT.format(value=date_str))

    raise ValueError(Errors.DATE_INVALID_TYPE)


def _generate_unique_short_name(base: str, max_len: int, existing: set[str]) -> str:
    """Генерирует уникальное короткое имя с суффиксом при коллизиях"""
    MAX_SUFFIX = 9999

    base = base[:max_len].rstrip()
    if base.lower() not in existing:
        return base

    suffix_len = 1
    while suffix_len <= MAX_SUFFIX:
        candidate = base[: max_len - len(str(suffix_len)) - 1] + f"_{suffix_len}"
        if candidate.lower() not in existing:
            return candidate
        suffix_len += 1

    raise ValueError(f"Не удалось сгенерировать уникальное имя для: {base}")


def _parse_gender(gender_raw: Any, row_num: int) -> str:
    """Парсинг и валидация пола с поддержкой разных написаний"""
    gender = str(gender_raw).strip() if gender_raw is not None else ""
    allowed_list = list(Limits.GENDER_MAP.keys())
    allowed_str = ", ".join(allowed_list)

    if not gender:
        raise ValueError(
            Errors.IMPORT_GENDER_REQUIRED.format(
                row_num=row_num,
                allowed=allowed_str,
            )
        )

    normalized = Limits.GENDER_MAP.get(gender.lower())
    if normalized is None:
        raise ValueError(
            Errors.IMPORT_GENDER_INVALID.format(
                row_num=row_num,
                gender=gender,
                allowed=allowed_str,
            )
        )
    return normalized


def _load_municipalities(names: set[str]) -> dict[str, Municipality]:
    """Загрузить муниципалитеты по набору имён"""
    return {
        m.name.lower(): m
        for m in Municipality.objects.annotate(name_lower=Lower("name")).filter(
            name_lower__in={n.lower() for n in names}
        )
    }


# =============================================================================
# Базовый ViewSet с экспортом и импортом
# =============================================================================


class BaseVSOSHViewSet(viewsets.ModelViewSet):
    """Базовый ViewSet для моделей VSOSH с поддержкой экспорта и импорта"""

    export_columns_map: dict[str, str] = {}
    import_required_fields: list[str] = []
    column_mapping: dict[str, str] = {}

    def get_export_queryset(self) -> models.QuerySet:
        """Получить queryset для экспорта с учетом фильтров"""
        return self.filter_queryset(self.get_queryset())

    def get_export_columns(self, requested_columns: list[str] | None = None) -> list[str]:
        """Получить список колонок для экспорта"""
        if requested_columns:
            return [col for col in requested_columns if col in self.export_columns_map]
        return list(self.export_columns_map.keys())

    def _normalize_headers(self, raw_headers: list) -> list[str | None]:
        """Преобразует русские заголовки в английские, дубликаты в None"""
        if not self.column_mapping:
            return [str(h).strip() if h else None for h in raw_headers]

        normalized: list[str | None] = []
        used_fields: set[str] = set()

        for header in raw_headers:
            if header is None:
                normalized.append(None)
                continue

            header_clean = str(header).strip()

            if header_clean in used_fields:
                normalized.append(None)
                continue

            if header_clean in self.column_mapping:
                field = self.column_mapping[header_clean]
                if field in used_fields:
                    normalized.append(None)
                else:
                    normalized.append(field)
                    used_fields.add(field)
            else:
                normalized.append(header_clean)
                used_fields.add(header_clean)

        return normalized

    def _get_reverse_mapping(self) -> dict[str, list[str]]:
        """Лениво создать и кэшировать обратный маппинг"""
        if not hasattr(self, "_reverse_mapping"):
            self._reverse_mapping: dict[str, list[str]] = {}
            for ru_header, en_field in self.column_mapping.items():
                self._reverse_mapping.setdefault(en_field, []).append(ru_header)
            for en_field in self._reverse_mapping:
                self._reverse_mapping[en_field].append(en_field)
        return self._reverse_mapping

    def _get_resolved_value(self, row_data: dict, field: str, default: Any = None) -> Any:
        """Получить значение поля из строки по нормализованному имени"""
        value = row_data.get(field)
        if value is not None:
            if not isinstance(value, str):
                value = str(value)
            value = value.strip()
            if value:
                return value
        return default

    def _validate_row(
        self,
        row_data: dict,
        field_limits: dict[str, int],
        row_num: int,
    ) -> None:
        """Проверить пустоту + длину. Только для колонок, реально присутствующих в файле"""
        reverse_mapping = self._get_reverse_mapping()

        for field, max_len in field_limits.items():
            # Пропускаем колонки, которых нет в этом файле
            if field not in row_data:
                continue

            value = row_data[field]
            display_names = reverse_mapping.get(field, [field])
            display_name = "/".join(display_names)

            # Пустая ячейка
            if value is None or (isinstance(value, str) and not value.strip()):
                raise ValueError(
                    Errors.EMPTY_CELL.format(row_num=row_num, display_name=display_name)
                )

            # Превышение длины
            if isinstance(value, str) and len(value) > max_len:
                raise ValueError(
                    Errors.IMPORT_FIELD_TOO_LONG.format(
                        row_num=row_num,
                        field=display_name,
                        max_len=max_len,
                        length=len(value),
                    )
                )

    @action(detail=False, methods=["get"], url_path="export")
    @audit_log(
        action="экспорт", model="Данные", get_new_values=_get_export_new_values, allowed_get=True
    )
    def export(self, request: Request) -> Response:
        """Экспорт данных в Excel с выбором колонок"""
        # === Этап 1: Валидация параметров запроса ===
        serializer = ExportRequestSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)

        columns = serializer.validated_data.get("columns")

        # === Этап 2: Получение данных с текущими фильтрами ===
        queryset = self.get_export_queryset()

        if not queryset.exists():
            return Response(
                {"detail": Errors.EXPORT_NO_DATA},
                status=status.HTTP_404_NOT_FOUND,
            )

        columns = self.get_export_columns(columns)

        # === Этап 3: Сериализация данных ===
        export_serializer = self.get_serializer(queryset, many=True)
        serialized_data = export_serializer.data

        # === Этап 4: Создание Excel-файла ===
        wb = Workbook()
        ws = wb.active
        if ws is None:
            return Response(
                {"detail": Errors.WORKSHEET_NOT_CREATED},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        ws.title = self.get_queryset().model._meta.verbose_name_plural

        # Заголовки
        headers = [self.export_columns_map[col] for col in columns]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        # Строки данных
        for row_num, row_data in enumerate(serialized_data, 2):
            for col_num, field in enumerate(columns, 1):
                value = row_data.get(field, "")
                ws.cell(row=row_num, column=col_num, value=value)

        # === Этап 5: Формирование ответа ===
        filename = f"{self.get_queryset().model._meta.verbose_name_plural}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        rows_count = len(serialized_data)
        cols_count = len(columns)

        # Сохраняем в request для последующего использования в аудите
        request._export_rows_count = rows_count  # type: ignore[attr-defined]

        response["X-Message"] = Messages.EXPORT_SUCCESS.format(
            rows_count=rows_count, cols_count=cols_count
        )
        response["Access-Control-Expose-Headers"] = "X-Message, Content-Disposition"
        return response  # type: ignore[return-value]

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser])
    @audit_log(
        action="импорт",
        model="Данные",
        get_new_values=_get_import_new_values,
    )
    def import_data(self, request: Request) -> Response:
        """Импорт данных из Excel/CSV файла с поддержкой русских заголовков"""
        # === Этап 1: Проверка файла ===
        if "file" not in request.FILES:
            return self._import_response(detail=Errors.IMPORT_FILE_REQUIRED)

        file = request.FILES["file"]

        file_ext = "." + file.name.rsplit(".", 1)[-1].lower() if "." in file.name else ""
        if file_ext not in Limits.SUPPORTED_IMPORT_FORMATS:
            return self._import_response(detail=Errors.IMPORT_INVALID_FORMAT)

        try:
            if file_ext == ".csv":
                try:
                    df = pd.read_csv(file, encoding="utf-8")
                except UnicodeDecodeError:
                    # Пробуем windows-1251
                    try:
                        file.seek(0)
                        df = pd.read_csv(file, encoding="windows-1251")
                    except Exception as e:
                        logger.warning(f"Ошибка чтения CSV (windows-1251): {e}")
                        return self._import_response(detail=Errors.IMPORT_CSV_ENCODING_ERROR)
            else:
                df = pd.read_excel(file, engine="openpyxl")
        except Exception as e:
            logger.warning(f"Ошибка чтения файла импорта: {e}")
            return self._import_response(detail=Errors.IMPORT_INVALID_FORMAT)

        # === Проверка заголовков файла на обязательные поля ===
        file_headers = list(df.columns)
        reverse_mapping = self._get_reverse_mapping()

        missing_headers = []
        for field in self.import_required_fields:
            possible_names = reverse_mapping.get(field, [field])
            has_any = any(name in file_headers for name in possible_names)
            if not has_any:
                missing_headers.append("/".join(possible_names))

        if missing_headers:
            error_msg = Errors.IMPORT_MISSING_HEADERS.format(fields=", ".join(missing_headers))
            return self._import_response(
                message=error_msg,
                error_count=1,
                error_details=[{"row": 1, "error": error_msg}],
            )

        # === Проверка лимита строк ===
        total_rows = len(df)
        if total_rows > Limits.IMPORT_MAX_ROWS:
            error_msg = Errors.IMPORT_MAX_ROWS_EXCEEDED.format(
                actual_rows=total_rows, max_rows=Limits.IMPORT_MAX_ROWS
            )
            return self._import_response(
                message=error_msg,
                error_count=1,
                error_details=[{"row": 1, "error": error_msg}],
            )

        # === Этап 2: Парсинг строк ===
        headers = self._normalize_headers(file_headers)

        rows_data = []
        error_details = []
        error_count = 0

        for row_num, (_, row) in enumerate(df.iterrows(), 2):
            row_data = {}
            for idx, header in enumerate(headers):
                if header:
                    val = row.iloc[idx]
                    row_data[header] = None if pd.isna(val) else val

            try:
                parsed = self._parse_import_row(row_data, row_num)
                rows_data.append(parsed)
            except Exception as e:
                error_count += 1
                error_details.append({"row": row_num, "error": str(e)})

        # === Этап 3: Валидация FK (собираем ВСЕ ошибки) ===
        fk_error_count, *fk_caches = self._validate_fk(rows_data, error_details)
        error_count += fk_error_count

        # === Этап 4: Возврат ошибок или запись в БД ===
        if error_details:
            return self._import_response(
                message=Messages.IMPORT_WITH_ERRORS.format(error_count=error_count),
                error_count=error_count,
                error_details=error_details,
            )

        with transaction.atomic():
            created_count, updated_count = self._bulk_process(rows_data, *fk_caches)

        # Инвалидация кэша фильтров после успешного импорта
        VSOSHCacheService.invalidate_filters()

        return self._import_response(
            message=Messages.IMPORT_SUCCESS.format(created=created_count, updated=updated_count),
            created=created_count,
            updated=updated_count,
        )

    def _import_response(
        self,
        message: str | None = None,
        detail: str | None = None,
        created: int | None = None,
        updated: int | None = None,
        error_count: int | None = None,
        error_details: list | None = None,
    ) -> Response:
        """Универсальный метод для формирования ответа импорта"""
        status_code = (
            status.HTTP_400_BAD_REQUEST if (detail or error_details) else status.HTTP_200_OK
        )

        data = {
            k: v
            for k, v in {
                "detail": detail,
                "message": message,
                "created": created,
                "updated": updated,
                "errors": error_count,
                "error_details": error_details or ([] if detail else None),
            }.items()
            if v is not None
        }

        return Response(data, status=status_code)

    def _validate_fk(self, rows_data: list[dict[str, Any]], error_details: list) -> tuple:
        """
        Валидация внешних ключей.
        Возвращает (error_count, *caches) — кэши передаются в _bulk_process.
        Переопределяется в подклассах.
        """
        return (0,)

    def _parse_import_row(self, row_data: dict[str, Any], row_num: int) -> dict[str, Any]:
        """Распарсить строку импорта. Должен быть переопределен в подклассе"""
        raise NotImplementedError("Необходимо реализовать в подклассе")

    def _bulk_process(self, rows_data: list[dict[str, Any]], *args) -> tuple[int, int]:
        """
        Пакетная обработка: создаёт и обновляет записи через bulk_create/bulk_update.
        Должен быть переопределен в подклассе.
        """
        raise NotImplementedError("Необходимо реализовать в подклассе")

    def perform_update(self, serializer):
        """Сохраняем старые данные ДО обновления для аудита"""
        self.request._audit_old_data = _model_to_dict_with_fk_ids(serializer.instance)
        super().perform_update(serializer)

    def perform_destroy(self, instance):
        """Проверка FK-зависимостей перед удалением + сохранение для аудита"""
        self.request._audit_old_data = _model_to_dict_with_fk_ids(instance)

        checks = {
            Municipality: ("education_institutions", Errors.MUNICIPALITY_IN_USE),
            Subject: ("olympiad_participations", Errors.SUBJECT_IN_USE),
            EducationInstitution: ("olympiad_participations", Errors.INSTITUTION_IN_USE),
            Participant: ("olympiad_participations", Errors.PARTICIPANT_IN_USE),
        }

        model_class = type(instance)
        if model_class in checks:
            related_name, error_msg = checks[model_class]
            if getattr(instance, related_name).exists():
                raise ValidationError(error_msg)

        instance.delete()
